"""
01_load_data.py — Chargement et nettoyage des données FactSet
Lit les .xlsx du dossier dataset/, extrait Date + prix, merge tout.
Sortie : data/processed/master_daily.pkl
"""
import pandas as pd
import numpy as np
import openpyxl
from pathlib import Path
import warnings
warnings.filterwarnings("ignore")

from config import (
    DATASET_DIR, DATA_PROCESSED, START_DATE, END_DATE,
    DAILY_FILES, MONTHLY_FILES
)


def parse_value(val):
    """Parse une valeur qui peut être en format européen (virgule décimale)."""
    if val is None or val == "-" or val == "":
        return np.nan
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        # Format européen : "6,45" → 6.45
        val = val.strip().replace(",", ".")
        try:
            return float(val)
        except ValueError:
            return np.nan
    return np.nan


def load_factset_daily(filename: str, price_col: str, name: str) -> pd.Series:
    """Charge un fichier FactSet (détecte automatiquement le format)."""
    filepath = DATASET_DIR / filename
    if not filepath.exists():
        print(f"  [SKIP] {filename} introuvable")
        return pd.Series(dtype=float, name=name)

    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Trouver la ligne header (contient 'Date')
    header_idx = None
    for i, row in enumerate(rows):
        if row and row[0] == "Date":
            header_idx = i
            break

    if header_idx is None:
        print(f"  [ERR] Pas de header 'Date' dans {filename}")
        return pd.Series(dtype=float, name=name)

    headers = [str(h) if h else "" for h in rows[header_idx]]
    data_rows = rows[header_idx + 1:]

    # Trouver la colonne prix
    col_idx = None
    price_col_lower = price_col.lower()
    for j, h in enumerate(headers):
        if price_col_lower in h.lower():
            col_idx = j
            break

    if col_idx is None:
        # Fallback : colonne 1 (souvent le prix principal)
        col_idx = 1
        print(f"  [WARN] '{price_col}' non trouvee dans {filename}, col {col_idx} ({headers[col_idx][:30]})")

    dates, values = [], []
    for row in data_rows:
        if row[0] is None:
            continue
        date_val = row[0]
        price_val = row[col_idx] if col_idx < len(row) else None

        if isinstance(date_val, str):
            try:
                date_val = pd.to_datetime(date_val)
            except:
                continue

        parsed = parse_value(price_val)
        if not np.isnan(parsed):
            values.append(parsed)
            dates.append(pd.Timestamp(date_val))

    if not dates:
        print(f"  [ERR] Aucune donnee valide dans {filename}")
        return pd.Series(dtype=float, name=name)

    s = pd.Series(values, index=pd.DatetimeIndex(dates), name=name)
    s = s.sort_index()
    s = s[~s.index.duplicated(keep="first")]
    print(f"  [OK] {name:12s} : {len(s):6d} obs, {s.index[0].date()} -> {s.index[-1].date()}")
    return s


def load_factset_monthly(filename: str, row_label: str, name: str) -> pd.Series:
    """Charge un fichier FactSet format horizontal (dates en colonnes)."""
    filepath = DATASET_DIR / filename
    if not filepath.exists():
        print(f"  [SKIP] {filename} introuvable")
        return pd.Series(dtype=float, name=name)

    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    date_row = None
    data_row = None
    for i, row in enumerate(rows):
        vals = [v for v in row if v is not None]
        if not vals:
            continue
        if date_row is None and isinstance(vals[0], str) and "/" in vals[0]:
            try:
                pd.to_datetime(vals[0])
                date_row = row
            except:
                pass
        if vals and str(vals[0]).strip().lower() == row_label.lower():
            data_row = row

    if date_row is None or data_row is None:
        print(f"  [ERR] Format non reconnu pour {filename}")
        return pd.Series(dtype=float, name=name)

    dates, values = [], []
    for j in range(1, min(len(date_row), len(data_row))):
        d = date_row[j]
        v = data_row[j]
        if d is None:
            continue
        parsed = parse_value(v)
        if not np.isnan(parsed):
            try:
                dates.append(pd.to_datetime(d))
                values.append(parsed)
            except:
                continue

    s = pd.Series(values, index=pd.DatetimeIndex(dates), name=name)
    s = s.sort_index()
    s = s[~s.index.duplicated(keep="first")]
    print(f"  [OK] {name:12s} : {len(s):6d} obs (mensuel), {s.index[0].date()} -> {s.index[-1].date()}")
    return s


def main():
    print("=" * 60)
    print("STEP 1 : CHARGEMENT DES DONNEES FACTSET")
    print("=" * 60)

    # ── Données quotidiennes ──
    print("\n--- Donnees quotidiennes ---")
    daily_series = {}
    for key, (fname, col, name) in DAILY_FILES.items():
        s = load_factset_daily(fname, col, name)
        if len(s) > 0:
            daily_series[name] = s

    df_daily = pd.DataFrame(daily_series)
    df_daily.index.name = "date"
    df_daily = df_daily.loc[START_DATE:END_DATE]
    df_daily = df_daily.ffill(limit=5)

    print(f"\n  Master daily : {df_daily.shape[0]} jours x {df_daily.shape[1]} colonnes")
    print(f"  Periode : {df_daily.index[0].date()} -> {df_daily.index[-1].date()}")
    print(f"\n  NaN restants :")
    nans = df_daily.isna().sum()
    for col in nans[nans > 0].index:
        pct = nans[col] / len(df_daily) * 100
        print(f"    {col:12s} : {nans[col]:5d} ({pct:.1f}%)")

    # ── Données mensuelles ──
    print("\n--- Donnees mensuelles ---")
    monthly_series = {}
    for key, (fname, label, name) in MONTHLY_FILES.items():
        s = load_factset_monthly(fname, label, name)
        if len(s) > 0:
            monthly_series[name] = s

    df_monthly = pd.DataFrame(monthly_series)
    df_monthly.index.name = "date"
    df_monthly = df_monthly.loc[START_DATE:END_DATE]

    if not df_monthly.empty:
        df_monthly_daily = df_monthly.reindex(df_daily.index, method="ffill")
        for col in df_monthly_daily.columns:
            df_daily[col] = df_monthly_daily[col]
        print(f"  Mensuelles ajoutees : {list(df_monthly_daily.columns)}")

    # ── Calcul du GSR ──
    print("\n--- Gold/Silver Ratio ---")
    df_daily["gsr"] = df_daily["gold"] / df_daily["silver"]
    valid_gsr = df_daily["gsr"].dropna()
    print(f"  GSR : {len(valid_gsr)} obs, Min={valid_gsr.min():.1f}, Max={valid_gsr.max():.1f}, "
          f"Mean={valid_gsr.mean():.1f}, Median={valid_gsr.median():.1f}")

    # ── Sauvegarde (pickle, pas de pyarrow nécessaire) ──
    out_path = DATA_PROCESSED / "master_daily.pkl"
    df_daily.to_pickle(out_path)
    # CSV backup
    df_daily.to_csv(DATA_PROCESSED / "master_daily.csv")
    print(f"\n  Sauvegarde : {out_path}")
    print(f"  Colonnes : {list(df_daily.columns)}")
    print(f"\n{'=' * 60}")
    print("STEP 1 TERMINE")
    print(f"{'=' * 60}")

    return df_daily


if __name__ == "__main__":
    df = main()
